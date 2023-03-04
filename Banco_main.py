from time import sleep
import sys
from PyQt5.QtWidgets import QMainWindow, QApplication
from PyQt5.QtWidgets import QWidget, QMessageBox
from bank import *
from login import *
from tela_transferir import *
from tela_cadastro import *
import openpyxl
import pandas as pd
import win32com.client as win32


class Tela_transferir(QMainWindow):
        def __init__(self):
                super().__init__()
                self.ui = Ui_tela_transferencia()
                self.ui.setupUi(self)
       


class Tela_banco(QMainWindow):
        def __init__(self):
                super().__init__()
                self.ui = Ui_DMBANK()
                self.ui.setupUi(self)
        
        def closeEvent(self, event):
                self.origem = Tela_login()
                self.origem.show()
                event.accept()


class Tela_cadastro(QMainWindow):
        def __init__(self):
                super().__init__()
                self.ui = Ui_cadastro_tela()
                self.ui.setupUi(self)
        def closeEvent(self, event):
                self.origem = Tela_login()
                self.origem.show()
                event.accept()

class Tela_login(QMainWindow):
        def __init__(self):
                super().__init__()
                self.ui = Ui_Form()
                self.ui.setupUi(self)
                self.tela_cadastro = Tela_cadastro()
                self.tela_banco = Tela_banco()
                self.tela_transferencia = Tela_transferir()
                self.ui.btn_Cadastrar.clicked.connect(self.abre_cadastro)
                self.tela_cadastro.ui.btn_enviar_cadastro.clicked.connect(self.envia_cadastro)
                self.ui.btn_login.clicked.connect(self.login)
                self.ui.btn_login.clicked.connect(self.saldo)
                self.ui.btn_login.clicked.connect(self.set_nome)
                self.ui.btn_login.clicked.connect(self.set_idade)
                #Tela principal do banco
                self.tela_banco.ui.btn_sacar.clicked.connect(self.sacar)
                self.tela_banco.ui.btn_depositar.clicked.connect(self.depositar)
                self.tela_banco.ui.btn_transferir.clicked.connect(self.abre_transferencia)
                self.tela_banco.ui.btn_transferir.clicked.connect(self.saldo)
                self.tela_banco.ui.label_conta_2.setText(" Conta-Corrente")
                #Tela de transferencia
                self.tela_transferencia.ui.label_conta_2.setText(" Conta-Corrente")
                self.tela_transferencia.ui.btn_transferir.clicked.connect(self.transferir)
                self.tela_transferencia.ui.btn_sair.clicked.connect(self.sair_transf)

        #Seta a idade.
        def set_idade(self):
                dados_login = self.ui.line_login.text()
                dados_senha = self.ui.line_senha.text()
                banco_dados = openpyxl.load_workbook("banco_dados.xlsx")
                pagina_bd = banco_dados['Planilha1']
                for linha in pagina_bd.iter_rows(min_row=1, max_row=10):
                        if linha[0].value == dados_login and linha[1].value == dados_senha:
                                valor = f" Idade: {linha[5].value} anos."
                                self.set_idade = str(valor)
                                self.tela_banco.ui.label_idade.setText(self.set_idade)
                                self.tela_transferencia.ui.label_idade.setText(self.set_idade)

        #Seta o nome.
        def set_nome(self):
                dados_login = self.ui.line_login.text()
                dados_senha = self.ui.line_senha.text()
                banco_dados = openpyxl.load_workbook("banco_dados.xlsx")
                pagina_bd = banco_dados['Planilha1']
                for linha in pagina_bd.iter_rows(min_row=1, max_row=10):
                        if linha[0].value == dados_login and linha[1].value == dados_senha:
                                valor = f" {linha[3].value} {linha[4].value}"
                                self.set_nome = str(valor)
                                self.tela_banco.ui.label_nome.setText(self.set_nome)
                                self.tela_transferencia.ui.label_nome.setText(self.set_nome)
                                
        #Seta o saldo.
        def saldo(self):
                dados_login = self.ui.line_login.text()
                dados_senha = self.ui.line_senha.text()
                banco_dados = openpyxl.load_workbook("banco_dados.xlsx")
                pagina_bd = banco_dados['Planilha1']
                for linha in pagina_bd.iter_rows(min_row=1, max_row=10):
                        if linha[0].value == dados_login and linha[1].value == dados_senha:
                                valor = float(linha[2].value)
                                
                                self.saldo = (f' Saldo: R${valor}')
                                self.tela_banco.ui.label_saldo.setText(self.saldo)
                                self.tela_transferencia.ui.label_saldo.setText(self.saldo)
        #função para efetuar saques da conta.
        def sacar(self):
                dados_login = self.ui.line_login.text()
                dados_senha = self.ui.line_senha.text()
                valor_sacar = self.tela_banco.ui.line_input.text()
                banco_dados = openpyxl.load_workbook("banco_dados.xlsx")
                pagina_bd = banco_dados['Planilha1']
                for linha in pagina_bd.iter_rows(min_row=1, max_row=10):
                        if linha[0].value == dados_login and linha[1].value == dados_senha:
                                valor = float(linha[2].value)
                                try:
                                        if not float(valor_sacar) > valor:
                                                valor -= float(valor_sacar)
                                                linha[2].value = valor
                                                banco_dados.save('banco_dados.xlsx')
                                                self.saldo = (f' Saldo: R${valor}')
                                                self.tela_banco.ui.label_TelaPrincipal.setText('     Saque conluida com sucesso.')
                                                self.tela_banco.ui.label_saldo.setText(self.saldo)
                                                self.tela_banco.ui.line_input.clear()
                                                
                                        else:
                                                self.tela_banco.ui.label_TelaPrincipal.setText(' Saldo insuficiente para concluir a transação.')
                                except:
                                        self.tela_banco.ui.label_TelaPrincipal.setText(' Comando desconhecido,\n por favor,\n digite, apenas números.')
                        
        def transferir(self):
                dados_login = self.ui.line_login.text()
                dados_senha = self.ui.line_senha.text()
                valor_transferir = self.tela_transferencia.ui.line_input_valor.text()
                destinatario = self.tela_transferencia.ui.line_input_destinatario.text()
                banco_dados = openpyxl.load_workbook("banco_dados.xlsx")
                pagina_bd = banco_dados['Planilha1']
                try:
                        for linha in pagina_bd.iter_rows(min_row=1):
                                if linha[0].value == dados_login and linha[1].value == dados_senha:
                                        valor = float(linha[2].value)
                                        for linha1 in pagina_bd.iter_rows(min_row=1):
                                                inf = True
                                                if linha1[0].value == destinatario:
                                                        inf = False
                                                        try:
                                                                if not float(valor_transferir) > valor:
                                                                        valor -= float(valor_transferir)
                                                                        linha[2].value = valor
                                                                        
                                                                        valor_transferido = float(valor_transferir)
                                                                        linha1[2].value += valor_transferido
                                                                        banco_dados.save('banco_dados.xlsx')


                                                                        self.saldo = (f' Saldo: R${valor}')
                                                                        self.tela_banco.ui.label_saldo.setText(self.saldo)
                                                                        self.tela_transferencia.ui.label_saldo.setText(self.saldo)
                                                                        self.tela_transferencia.ui.line_input_valor.clear()
                                                                        self.tela_transferencia.ui.line_input_destinatario.clear()

                                                                        QMessageBox.about(self.tela_transferencia, 'SUCESSO', f'R${valor_transferido} transferido\n para {destinatario}.')
                                                                
                                                                else:
                                                                        QMessageBox.about(self.tela_transferencia, 'AVISO', 'Saldo insuficiente.')
                                        
                                                        except:
                                                                QMessageBox.about(self.tela_transferencia, 'AVISO', 'Valor invalido')
                                        if inf == True:
                                                QMessageBox.about(self.tela_transferencia, 'AVISO', 'Usuário Não encontrado.')
                except:
                        QMessageBox.about(self.tela_transferencia, 'AVISO', 'Valor invalido')

        #função para efetuar depósitos na conta
        def depositar(self):
                dados_login = self.ui.line_login.text()
                dados_senha = self.ui.line_senha.text()
                valor_depositar = self.tela_banco.ui.line_input.text()
                banco_dados = openpyxl.load_workbook("banco_dados.xlsx")
                pagina_bd = banco_dados['Planilha1']
                for linha in pagina_bd.iter_rows(min_row=1, max_row=10):
                        if linha[0].value == dados_login and linha[1].value == dados_senha:
                                valor = float(linha[2].value)
                                limite_deposito = float(100000)
                                try:
                                        if not float(valor_depositar) > limite_deposito:
                                                valor += float(valor_depositar)
                                                linha[2].value = valor
                                                banco_dados.save('banco_dados.xlsx')
                                                self.saldo = (f' Saldo: R${valor}')
                                                self.tela_banco.ui.label_TelaPrincipal.setText('     Depósito conluida com sucesso.')
                                                self.tela_banco.ui.label_saldo.setText(self.saldo)
                                                self.tela_banco.ui.line_input.clear()
                                        
                                        else:
                                                self.tela_banco.ui.label_TelaPrincipal.setText(' Limite de depósito R$100.000,00,\n ultrapassado,\n por favor compareça a uma agencia..')
                                except:
                                        self.tela_banco.ui.label_TelaPrincipal.setText(' Comando desconhecido,\n por favor,\n digite, apenas números.')
                                                      
                          
        def login(self): #Verifica dados e efetua login.
                dados_login = self.ui.line_login.text()
                dados_senha = self.ui.line_senha.text()
                banco_dados = openpyxl.load_workbook("banco_dados.xlsx")
                pagina_bd = banco_dados['Planilha1']
                set1 = True
                for linha in pagina_bd.iter_rows(min_row=1, max_row=10):
                        if linha[0].value == dados_login:
                                if  linha[1].value == dados_senha:
                                        set1 = False
                                        self.abre_banco()
                                        break
                                else:
                                        QMessageBox.about(w, 'ERRO', 'Senha inválida')
                                        set1 = False
                                        break
                if set1 == True:        
                        QMessageBox.about(w, 'ERRO', 'Usuário incorreto\n           ou\n não cadastrado.')
                        

        def abre_transferencia(self): #Abre tela de transferencia.
                self.tela_transferencia.show()
                        
                                
        def abre_cadastro(self): #Abre tela de cadastro.
                self.tela_cadastro.show()
                self.hide()
                
        def envia_cadastro(self): #Cadastra o usuario.
                nome = self.tela_cadastro.ui.line_nome.text()
                sobrenome = self.tela_cadastro.ui.line_sobrenome.text()
                idade = self.tela_cadastro.ui.line_idade.text()
                loguin_cadastrar = self.tela_cadastro.ui.line_cadastrar_login.text()
                senha_cadastrar = self.tela_cadastro.ui.line_cadastrar_senha.text()
                saldo = float(0)
                banco_dados = openpyxl.load_workbook("banco_dados.xlsx")
                pagina_bd = banco_dados['Planilha1']
                inf = False
                try:
                        if not idade == '':
                                idade_test = int(idade)
                
                except:
                        self.tela_cadastro.ui.label_2.setText(' ERRO: Usuário deve ter no minimo 18 anos.')
                        inf = True

                try:
                        if len(nome) < 3:
                                self.tela_cadastro.ui.label_2.setText(' ERRO: Nome deve conter mais de 3 letras.')
                                inf = True

                        elif len(sobrenome) < 3:
                                self.tela_cadastro.ui.label_2.setText(' ERRO: Sobrenome deve conter mais de 3 letras.')
                                inf = True

                        elif idade == '':
                                self.tela_cadastro.ui.label_2.setText(' ERRO: Usuário deve ter no minimo 18 anos.')
                                inf = True

                        elif idade_test < 18:
                                self.tela_cadastro.ui.label_2.setText(' ERRO: Usuário deve ter no minimo 18 anos.')
                                inf = True
                        elif idade_test > 120:
                                self.tela_cadastro.ui.label_2.setText(' ERRO: Usuário deve ter no máximo 120 anos.')
                                inf = True
                        
                        elif len(loguin_cadastrar) < 5:
                                self.tela_cadastro.ui.label_2.setText(' ERRO: Login deve conter mais de 5 letras.')
                                inf = True
                        
                        elif len(senha_cadastrar) < 5:
                                self.tela_cadastro.ui.label_2.setText(' ERRO: Senha deve conter mais de 5 digitos.')
                                inf = True

                        else:
                                for linha in pagina_bd.iter_rows(min_row=2):
                                        if linha[0].value == str(loguin_cadastrar):
                                                self.tela_cadastro.ui.label_2.setText(' ERRO: Usuário já cadastrado.')
                                                inf = True
                                                break
                                        else: 
                                                inf = False


                                if inf == False:
                                        pagina_bd.append([loguin_cadastrar, senha_cadastrar, saldo, nome, sobrenome, idade,])
                                        banco_dados.save('banco_dados.xlsx')
                                        self.tela_cadastro.ui.label_2.setText('Usuário cadastrado com sucesso.\n Feche esta janela, para fazer login')
                                        QMessageBox.about(self.tela_cadastro, 'AVISO', 'Usuário cadastrado com sucesso.')
                                        
                except:
                        self.tela_cadastro.ui.label_2.setText(' ERRO: Algo deu errado.')
                        QMessageBox.about(self.tela_cadastro, 'ERRO', 'Algo deu errado..')


        def sair_transf(self):
                self.tela_transferencia.close()      

        def abre_banco(self): #Abre a tela do Banco.
                self.tela_banco.show()
                self.hide()



if __name__ == "__main__":
        app = QApplication(sys.argv)
        w = Tela_login()
        w.show()
        sys.exit(app.exec_())